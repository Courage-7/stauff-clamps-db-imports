#!/usr/bin/env python3
"""Generate SQL inserts for the STAUFF Clamps workbook.

This importer keeps the database shape to seven tables: one table per sheet.
The table names are the workbook/overview catalogue labels.
It does not connect to Postgres; it writes a reviewable SQL data file.
"""

from __future__ import annotations

import argparse
from pathlib import Path
from typing import Any, Iterable

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DEFAULT_WORKBOOK = ROOT / "STAUFF_pages_45_50_51_52_53_ordering_codes - Copy.xlsx"
DEFAULT_OUTPUT = ROOT / "db" / "stauff_clamps_data.sql"

COMMON_COLUMNS = [
    "source_file",
    "source_sheet",
    "source_row",
    "catalogue_page_title",
    "catalogue_name",
    "permutation_summary",
]

PAGE_COLUMNS = [
    "ordering_code",
    "product_type",
    "product_description",
    "stauff_groups",
    "din_groups",
    "thread_code",
    "thread",
    "size_d1",
    "modifier",
    "material_code",
    "material_description",
    "notes",
]

SHEET_CONFIG = {
    "Overview": {
        "table": "Catalogue Overview",
        "data_start_row": 6,
        "metadata": {"title": 1, "name": 2, "permutations": 3, "section": 4},
        "columns": COMMON_COLUMNS
        + ["section_title", "page", "products", "product_type", "permutations", "sheet"],
        "source_columns": ["page", "products", "product_type", "permutations", "sheet"],
    },
    "Page45": {
        "table": "Elongated Weld Plates (SPAL-DUEB, SPAS-DUEB)",
        "data_start_row": 6,
        "metadata": {"title": 1, "name": 2, "permutations": 3},
        "columns": COMMON_COLUMNS + PAGE_COLUMNS,
        "source_columns": PAGE_COLUMNS,
    },
    "Page50": {
        "table": "Cover Plates (DPAL, DPAS)",
        "data_start_row": 6,
        "metadata": {"title": 1, "name": 2, "permutations": 3},
        "columns": COMMON_COLUMNS + PAGE_COLUMNS,
        "source_columns": PAGE_COLUMNS,
    },
    "Page51": {
        "table": "Bolts and Screws (AS, IS)",
        "data_start_row": 6,
        "metadata": {"title": 1, "name": 2, "permutations": 3},
        "columns": COMMON_COLUMNS + PAGE_COLUMNS,
        "source_columns": PAGE_COLUMNS,
    },
    "Page52": {
        "table": "Safety Washers (SI DIN 93, SI DIN 463)",
        "data_start_row": 6,
        "metadata": {"title": 1, "name": 2, "permutations": 3},
        "columns": COMMON_COLUMNS + PAGE_COLUMNS,
        "source_columns": PAGE_COLUMNS,
    },
    "Page53": {
        "table": "Safety Locking Plate & Stacking Bolt (SIP, AF)",
        "data_start_row": 6,
        "metadata": {"title": 1, "name": 2, "permutations": 3},
        "columns": COMMON_COLUMNS + PAGE_COLUMNS,
        "source_columns": PAGE_COLUMNS,
    },
    "Materials": {
        "table": "Material code reference",
        "data_start_row": 5,
        "metadata": {"title": 1, "name": 2, "permutations": 3},
        "columns": COMMON_COLUMNS + ["code", "material", "used_on_products"],
        "source_columns": ["code", "material", "used_on_products"],
    },
}

EXPECTED_COUNTS = {
    "Catalogue Overview": 25,
    "Elongated Weld Plates (SPAL-DUEB, SPAS-DUEB)": 200,
    "Cover Plates (DPAL, DPAS)": 103,
    "Bolts and Screws (AS, IS)": 112,
    "Safety Washers (SI DIN 93, SI DIN 463)": 24,
    "Safety Locking Plate & Stacking Bolt (SIP, AF)": 120,
    "Material code reference": 6,
}


def first_cell_text(ws: Any, row_number: int) -> str | None:
    value = ws.cell(row=row_number, column=1).value
    return to_text(value)


def to_text(value: Any) -> str | None:
    if value is None:
        return None
    text = str(value)
    return text if text != "" else None


def row_has_data(values: Iterable[Any]) -> bool:
    return any(to_text(value) is not None for value in values)


def sql_literal(value: str | None) -> str:
    if value is None:
        return "NULL"
    return "'" + value.replace("'", "''") + "'"


def sql_identifier(name: str) -> str:
    return '"' + name.replace('"', '""') + '"'


def qualified_table_name(table: str) -> str:
    return f"stauff_clamps.{sql_identifier(table)}"


def build_rows(workbook_path: Path) -> dict[str, list[dict[str, str | None]]]:
    wb = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    output: dict[str, list[dict[str, str | None]]] = {}

    for sheet_name, config in SHEET_CONFIG.items():
        ws = wb[sheet_name]
        metadata = config["metadata"]
        base_values = {
            "source_file": workbook_path.name,
            "source_sheet": sheet_name,
            "catalogue_page_title": first_cell_text(ws, metadata["title"]),
            "catalogue_name": first_cell_text(ws, metadata["name"]),
            "permutation_summary": first_cell_text(ws, metadata["permutations"]),
        }
        if "section" in metadata:
            base_values["section_title"] = first_cell_text(ws, metadata["section"])

        rows: list[dict[str, str | None]] = []
        for row_number, row in enumerate(
            ws.iter_rows(
                min_row=config["data_start_row"],
                max_row=ws.max_row,
                values_only=True,
            ),
            start=config["data_start_row"],
        ):
            if not row_has_data(row):
                continue

            record = dict(base_values)
            record["source_row"] = str(row_number)
            for index, column_name in enumerate(config["source_columns"]):
                record[column_name] = to_text(row[index] if index < len(row) else None)
            rows.append(record)

        output[config["table"]] = rows

    return output


def render_insert(table: str, columns: list[str], rows: list[dict[str, str | None]]) -> str:
    if not rows:
        return f"-- No rows for {qualified_table_name(table)}\n"

    column_sql = ", ".join(columns)
    values_sql = []
    for row in rows:
        values = ", ".join(sql_literal(row.get(column)) for column in columns)
        values_sql.append(f"  ({values})")
    return (
        f"insert into {qualified_table_name(table)} ({column_sql}) values\n"
        + ",\n".join(values_sql)
        + ";\n"
    )


def render_sql(rows_by_table: dict[str, list[dict[str, str | None]]]) -> str:
    table_names = [config["table"] for config in SHEET_CONFIG.values()]
    lines = [
        "-- Generated by scripts/import_stauff_clamps.py",
        "-- Run after db/stauff_clamps_schema.sql.",
        "begin;",
        "truncate table "
        + ", ".join(qualified_table_name(table_name) for table_name in table_names)
        + " restart identity;",
        "",
    ]

    for config in SHEET_CONFIG.values():
        table = config["table"]
        lines.append(render_insert(table, config["columns"], rows_by_table[table]))

    lines.append("commit;\n")
    return "\n".join(lines)


def validate_counts(rows_by_table: dict[str, list[dict[str, str | None]]]) -> None:
    actual_counts = {table: len(rows) for table, rows in rows_by_table.items()}
    errors = []
    for table, expected in EXPECTED_COUNTS.items():
        actual = actual_counts.get(table)
        if actual != expected:
            errors.append(f"{table}: expected {expected}, got {actual}")
    if errors:
        raise SystemExit("Row-count validation failed:\n" + "\n".join(errors))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate STAUFF Clamps SQL inserts.")
    parser.add_argument("--workbook", type=Path, default=DEFAULT_WORKBOOK)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--dry-run", action="store_true", help="Validate and print counts only.")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    workbook_path = args.workbook.resolve()
    if not workbook_path.exists():
        raise SystemExit(f"Workbook not found: {workbook_path}")

    rows_by_table = build_rows(workbook_path)
    validate_counts(rows_by_table)

    for table, rows in rows_by_table.items():
        print(f"{table}: {len(rows)} rows")

    if args.dry_run:
        return

    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(render_sql(rows_by_table), encoding="utf-8")
    print(f"Wrote {args.output}")


if __name__ == "__main__":
    main()
